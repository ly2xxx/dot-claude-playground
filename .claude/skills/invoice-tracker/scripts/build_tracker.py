#!/usr/bin/env python3
"""Deterministic invoice cost-tracker and anomaly detector.

Reads two CSVs produced from the factor/service-charge invoices:
  - line_items.csv : one row per charge line
  - invoices.csv   : one row per invoice (footer totals + period)

Writes:
  - cost_tracker.xlsx            : multi-sheet workbook for tax tracking
  - anomaly_report_<invoice>.md  : one point-in-time report per invoice, each
                                   compared only against earlier invoices

All money is treated as the owner's apportioned, VAT-inclusive cost
(the "your_cost" column) unless stated otherwise. No figures are invented:
the script only aggregates and compares what is in the CSVs.

Usage:
  python build_tracker.py --input-dir DIR [--output-dir DIR]
                          [--tax-start-day 6] [--tax-start-month 4]
                          [--high-ratio 1.5] [--low-ratio 0.5]
                          [--material 15]
"""

from __future__ import annotations

import argparse
import csv
import re
import statistics
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATE_RE = re.compile(r"(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})")
MONEY_FMT = "#,##0.00"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
ANOMALY_FILL = PatternFill("solid", fgColor="FFF2CC")
CREDIT_FILL = PatternFill("solid", fgColor="E2EFDA")
SEVERITY_FILL = {
    "high": PatternFill("solid", fgColor="F4B0B0"),
    "medium": PatternFill("solid", fgColor="FCE4A6"),
    "low": PatternFill("solid", fgColor="E2EFDA"),
}


# --------------------------------------------------------------------------
# Parsing helpers
# --------------------------------------------------------------------------
def parse_date(value: str):
    value = (value or "").strip()
    if not value:
        return None
    return datetime.strptime(value, "%d/%m/%Y").date()


def fmt_date(d) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


def parse_money(value: str) -> float:
    value = (value or "").strip().replace(",", "")
    if not value:
        return 0.0
    return float(value)


def parse_service_period(description: str):
    """Return (from_date, to_date) parsed from a 'dd/mm/yyyy - dd/mm/yyyy'
    range embedded in the description, or (None, None)."""
    m = DATE_RE.search(description or "")
    if not m:
        return None, None
    try:
        return (
            datetime.strptime(m.group(1), "%d/%m/%Y").date(),
            datetime.strptime(m.group(2), "%d/%m/%Y").date(),
        )
    except ValueError:
        return None, None


def norm_desc(description: str) -> str:
    """Collapse a description to a stable signature: drop dates, digits and
    punctuation so recurring lines match across quarters."""
    text = DATE_RE.sub(" ", description or "")
    text = re.sub(r"[0-9]", " ", text)
    text = re.sub(r"[^a-z ]", " ", text.lower())
    return re.sub(r"\s+", " ", text).strip()


def tax_year(d, start_day: int, start_month: int) -> str:
    """UK-style tax-year label, e.g. 2025/26 for 06/04/2025-05/04/2026."""
    if d is None:
        return "(undated)"
    start = (d.month, d.day) >= (start_month, start_day)
    y = d.year if start else d.year - 1
    return f"{y}/{str(y + 1)[-2:]}"


# --------------------------------------------------------------------------
# Loading
# --------------------------------------------------------------------------
def load_invoices(path: Path) -> dict:
    invoices = {}
    with path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            no = row["invoice_no"].strip()
            invoices[no] = {
                "invoice_no": no,
                "invoice_date": parse_date(row["invoice_date"]),
                "period_from": parse_date(row["period_from"]),
                "period_to": parse_date(row["period_to"]),
                "subtotal": parse_money(row.get("subtotal", "")),
                "vat": parse_money(row.get("vat", "")),
                "invoice_total": parse_money(row.get("invoice_total", "")),
                "balance_bf": parse_money(row.get("balance_bf", "")),
                "balance": parse_money(row.get("balance", "")),
            }
    return invoices


def load_items(path: Path) -> list:
    items = []
    with path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            desc = row["description"].strip()
            sf, st = parse_service_period(desc)
            items.append(
                {
                    "invoice_no": row["invoice_no"].strip(),
                    "invoice_date": parse_date(row["invoice_date"]),
                    "period_from": parse_date(row["period_from"]),
                    "period_to": parse_date(row["period_to"]),
                    "category": row["category"].strip(),
                    "charge_date": parse_date(row["charge_date"]),
                    "vendor": row["vendor"].strip(),
                    "description": desc,
                    "total_amount": parse_money(row["total_amount"]),
                    "share": row["share"].strip(),
                    "your_cost": parse_money(row["your_cost"]),
                    "service_from": sf,
                    "service_to": st,
                    "desc_sig": norm_desc(desc),
                }
            )
    return items


# --------------------------------------------------------------------------
# Anomaly detection
# --------------------------------------------------------------------------
def detect_anomalies(items, invoices, order, cfg) -> list:
    """Findings attributable to the current invoice (order[-1]).

    `order` lists every invoice visible at this point, oldest -> newest, and
    `items` must contain only those invoices' lines. Reconciliation, per-line
    and duplicate checks apply to the current invoice; drift checks compare it
    against the earlier ones. Call once per invoice (each as the newest) to
    build per-invoice, point-in-time reports."""
    findings = []

    def add(atype, severity, inv, detail, vendor="", category="", amount=None):
        findings.append(
            {
                "type": atype,
                "severity": severity,
                "invoice_no": inv,
                "category": category,
                "vendor": vendor,
                "amount": amount,
                "detail": detail,
            }
        )

    by_inv = defaultdict(list)
    for it in items:
        by_inv[it["invoice_no"]].append(it)
    current = order[-1]
    priors = order[:-1]
    cur_items = by_inv[current]

    # ---- Reconciliation: sum(your_cost) vs invoice footer total ----
    stated = invoices.get(current, {}).get("invoice_total")
    total = round(sum(i["your_cost"] for i in cur_items), 2)
    if stated is not None and abs(total - stated) > 0.01:
        add(
            "Reconciliation mismatch",
            "high",
            current,
            f"Line items sum to {total:.2f} but invoice total is {stated:.2f} "
            f"(diff {total - stated:+.2f}).",
            amount=total - stated,
        )

    # ---- Per-line date checks (current invoice) ----
    for it in cur_items:
        pf, pt, cd = it["period_from"], it["period_to"], it["charge_date"]
        if cd and pf and pt and not (pf <= cd <= pt):
            if cd < pf:
                gap = (pf - cd).days
                direction = f"{gap} days before period start"
            else:
                gap = (cd - pt).days
                direction = f"{gap} days after period end"
            sev = "high" if gap > 180 else "medium" if gap > 31 else "low"
            add(
                "Charge date outside invoice period",
                sev,
                it["invoice_no"],
                f"Charged {fmt_date(cd)} ({direction}); period "
                f"{fmt_date(pf)}-{fmt_date(pt)}; £{it['your_cost']:.2f}. \"{it['description'][:60]}\"",
                it["vendor"],
                it["category"],
                it["your_cost"],
            )
        # Service period (from the description) wholly outside the quarter
        sf, st = it["service_from"], it["service_to"]
        if sf and st and pf and pt:
            if st < pf:
                gap = (pf - st).days
                sev = "high" if gap > 180 else "medium" if gap > 31 else "low"
                add(
                    "Stale service period",
                    sev,
                    it["invoice_no"],
                    f"Service period {fmt_date(sf)}-{fmt_date(st)} ends "
                    f"{gap} days before the invoice quarter starts "
                    f"({fmt_date(pf)}); £{it['your_cost']:.2f}. \"{it['description'][:50]}\"",
                    it["vendor"],
                    it["category"],
                    it["your_cost"],
                )

    # ---- Credits / negative lines (informational, current invoice) ----
    for it in cur_items:
        if it["your_cost"] < 0:
            add(
                "Credit / negative line",
                "low",
                it["invoice_no"],
                f"{it['your_cost']:.2f} — \"{it['description'][:60]}\"",
                it["vendor"],
                it["category"],
                it["your_cost"],
            )

    # ---- Duplicate within the current invoice (same vendor+desc+amount+date) ----
    seen = defaultdict(list)
    for it in cur_items:
        key = (it["vendor"], it["description"], it["your_cost"], it["charge_date"])
        seen[key].append(it)
    for key, group in seen.items():
        if len(group) > 1:
            add(
                "Possible duplicate line (same invoice)",
                "high",
                current,
                f"{len(group)}x identical lines: {fmt_date(key[3])} "
                f"{key[0]} {key[2]:.2f} \"{key[1][:50]}\"",
                key[0],
                group[0]["category"],
                key[2],
            )

    # ---- Near-duplicate within invoice (same vendor+service period+cost, different description) ----
    svc_groups = defaultdict(list)
    for it in cur_items:
        sf, st = it["service_from"], it["service_to"]
        if sf and st:
            key = (it["vendor"], sf, st, round(it["your_cost"], 2))
            svc_groups[key].append(it)
    for key, group in svc_groups.items():
        if len(group) > 1 and len({it["description"] for it in group}) > 1:
            add(
                "Possible duplicate line (same invoice)",
                "high",
                current,
                f"{len(group)}x same vendor/period/cost with different descriptions: "
                f"{key[0]} {fmt_date(key[1])}-{fmt_date(key[2])} {key[3]:.2f} each — "
                + " | ".join(f"\"{it['description'][:40]}\"" for it in group),
                key[0],
                group[0]["category"],
                key[3] * len(group),
            )

    if not priors:
        return findings

    # ---- Cross-invoice duplicate: same vendor + overlapping service period ----
    prior_set = set(priors)
    prior_items_list = [it for it in items if it["invoice_no"] in prior_set]
    seen_cross: set = set()
    for it in cur_items:
        sf, st = it["service_from"], it["service_to"]
        if not sf or not st:
            continue
        for pit in prior_items_list:
            psf, pst = pit["service_from"], pit["service_to"]
            if not psf or not pst:
                continue
            if it["vendor"] != pit["vendor"]:
                continue
            if sf > pst or st < psf:
                continue  # no overlap
            overlap_days = (min(st, pst) - max(sf, psf)).days + 1
            shorter = min((st - sf).days + 1, (pst - psf).days + 1)
            if overlap_days / shorter < 0.5:
                continue  # consecutive billing periods — boundary overlap, not a duplicate
            dedup_key = (it["vendor"], sf, st, pit["invoice_no"], psf, pst)
            if dedup_key in seen_cross:
                continue
            seen_cross.add(dedup_key)
            same_amt = abs(it["your_cost"] - pit["your_cost"]) < 0.01
            sev = "high" if same_amt else "medium"
            add(
                "Cross-invoice duplicate charge",
                sev,
                current,
                f"{it['vendor']}: service {fmt_date(sf)}-{fmt_date(st)} "
                f"overlaps {overlap_days}d with inv {pit['invoice_no']} "
                f"{fmt_date(psf)}-{fmt_date(pst)}; £{it['your_cost']:.2f}. "
                f"\"{it['description'][:40]}\" vs \"{pit['description'][:40]}\"",
                it["vendor"],
                it["category"],
                it["your_cost"],
            )

    # ---- Vendor drift (current vs all priors) ----
    vendors_by_inv = {no: {i["vendor"] for i in by_inv[no]} for no in order}
    prior_vendors = set().union(*(vendors_by_inv[no] for no in priors))
    cur_vendors = vendors_by_inv[current]
    for v in sorted(cur_vendors - prior_vendors):
        amt = round(sum(i["your_cost"] for i in cur_items if i["vendor"] == v), 2)
        add(
            "New vendor",
            "medium",
            current,
            f"'{v}' appears this quarter but not in any prior quarter "
            f"(cost {amt:.2f}).",
            v,
            amount=amt,
        )
    for v in sorted(prior_vendors - cur_vendors):
        n_prior = sum(1 for no in priors if v in vendors_by_inv[no])
        if n_prior < 2:
            continue  # one-off contractors are not a "dropped" recurring vendor
        sev = "medium" if n_prior == len(priors) else "low"
        add(
            "Dropped vendor",
            sev,
            current,
            f"'{v}' billed in {n_prior}/{len(priors)} prior quarters but is "
            f"absent this quarter.",
            v,
        )

    # ---- Recurring item missing (description signature; category-agnostic so a
    # line the factor moved between section headers is not flagged as missing) ----
    sig_in_inv = defaultdict(set)
    sig_category = {}
    for it in items:
        ds = it["desc_sig"]
        if not ds:
            continue
        sig_in_inv[ds].add(it["invoice_no"])
        sig_category.setdefault(ds, it["category"])
    cur_sigs = {i["desc_sig"] for i in cur_items}
    for ds, invs in sorted(sig_in_inv.items()):
        prior_hits = [no for no in priors if no in invs]
        if len(prior_hits) == len(priors) and ds not in cur_sigs:
            add(
                "Recurring item missing",
                "medium",
                current,
                f"'{ds}' (e.g. {sig_category[ds]}) appeared in all {len(priors)} "
                f"prior quarters but not this quarter.",
                category=sig_category[ds],
            )

    # ---- Abnormal category totals (current vs prior mean) ----
    cat_totals = defaultdict(dict)  # category -> inv_no -> total
    for it in items:
        cat_totals[it["category"]][it["invoice_no"]] = (
            cat_totals[it["category"]].get(it["invoice_no"], 0.0) + it["your_cost"]
        )
    for cat, per_inv in sorted(cat_totals.items()):
        prior_vals = [round(per_inv[no], 2) for no in priors if no in per_inv]
        cur_val = round(per_inv.get(current, 0.0), 2)
        if not prior_vals:
            continue
        mean_prior = statistics.mean(prior_vals)
        if mean_prior <= 0:
            continue
        if current not in per_inv:
            continue  # handled by recurring/category-dropped logic
        ratio = cur_val / mean_prior
        if abs(cur_val - mean_prior) >= cfg.material and (
            ratio >= cfg.high_ratio or ratio <= cfg.low_ratio
        ):
            direction = "higher" if ratio >= cfg.high_ratio else "lower"
            add(
                "Abnormal category total",
                "medium" if direction == "higher" else "low",
                current,
                f"{cat}: {cur_val:.2f} this quarter vs prior average "
                f"{mean_prior:.2f} ({ratio:.1f}x, {direction}).",
                category=cat,
                amount=cur_val - mean_prior,
            )

    # ---- Abnormal vendor totals (current vs prior mean) ----
    ven_totals = defaultdict(dict)
    for it in items:
        ven_totals[it["vendor"]][it["invoice_no"]] = (
            ven_totals[it["vendor"]].get(it["invoice_no"], 0.0) + it["your_cost"]
        )
    for ven, per_inv in sorted(ven_totals.items()):
        prior_vals = [round(per_inv[no], 2) for no in priors if no in per_inv]
        if len(prior_vals) < 2 or current not in per_inv:
            continue
        mean_prior = statistics.mean(prior_vals)
        if mean_prior <= 0:
            continue
        cur_val = round(per_inv[current], 2)
        ratio = cur_val / mean_prior
        if abs(cur_val - mean_prior) >= cfg.material and (
            ratio >= cfg.high_ratio or ratio <= cfg.low_ratio
        ):
            direction = "higher" if ratio >= cfg.high_ratio else "lower"
            add(
                "Abnormal vendor total",
                "low",
                current,
                f"{ven}: {cur_val:.2f} this quarter vs prior average "
                f"{mean_prior:.2f} ({ratio:.1f}x, {direction}).",
                ven,
                amount=cur_val - mean_prior,
            )

    return findings


# --------------------------------------------------------------------------
# Workbook
# --------------------------------------------------------------------------
def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _autosize(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def build_workbook(items, invoices, order, findings, cfg, out_path: Path):
    wb = Workbook()

    # Flags per (invoice, description, charge_date) for the line sheet
    flagged = defaultdict(list)
    for f in findings:
        flagged[f["invoice_no"]].append(f)

    # ---------- Summary ----------
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Service-charge cost tracker"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Invoice", "Invoice date", "Period from", "Period to",
               "Sub-total", "VAT", "Invoice total", "Line-items sum", "Reconciles?"])
    _style_header(ws, 3, 9)
    grand = 0.0
    by_inv = defaultdict(list)
    for it in items:
        by_inv[it["invoice_no"]].append(it)
    for no in order:
        m = invoices.get(no, {})
        s = round(sum(i["your_cost"] for i in by_inv[no]), 2)
        grand += s
        ok = "yes" if m.get("invoice_total") is not None and abs(
            s - m["invoice_total"]) <= 0.01 else "CHECK"
        ws.append([no, fmt_date(m.get("invoice_date")), fmt_date(m.get("period_from")),
                   fmt_date(m.get("period_to")), m.get("subtotal"), m.get("vat"),
                   m.get("invoice_total"), s, ok])
    trow = ws.max_row + 1
    ws.cell(row=trow, column=1, value="TOTAL")
    ws.cell(row=trow, column=8, value=round(grand, 2))
    for c in range(1, 10):
        ws.cell(row=trow, column=c).fill = TOTAL_FILL
        ws.cell(row=trow, column=c).font = Font(bold=True)
    for r in range(4, trow + 1):
        for c in (5, 6, 7, 8):
            ws.cell(row=r, column=c).number_format = MONEY_FMT
    _autosize(ws, [12, 13, 13, 13, 12, 10, 13, 14, 11])

    # tax-year totals block
    ty_totals = defaultdict(float)
    for it in items:
        ty_totals[tax_year(it["charge_date"], cfg.tax_start_day,
                           cfg.tax_start_month)] += it["your_cost"]
    r0 = trow + 3
    ws.cell(row=r0, column=1, value="Totals by tax year (by charge date)").font = Font(bold=True)
    ws.cell(row=r0 + 1, column=1, value="Tax year").fill = HEADER_FILL
    ws.cell(row=r0 + 1, column=1).font = HEADER_FONT
    ws.cell(row=r0 + 1, column=2, value="Total cost").fill = HEADER_FILL
    ws.cell(row=r0 + 1, column=2).font = HEADER_FONT
    for i, ty in enumerate(sorted(ty_totals), start=1):
        ws.cell(row=r0 + 1 + i, column=1, value=ty)
        c = ws.cell(row=r0 + 1 + i, column=2, value=round(ty_totals[ty], 2))
        c.number_format = MONEY_FMT
    ws.cell(row=r0, column=4,
            value="Tax year = UK 6 Apr-5 Apr by default; see Line Items 'Tax year' column.")

    ws.freeze_panes = "A4"

    # ---------- Line Items ----------
    ws = wb.create_sheet("Line Items")
    cols = ["Tax year", "Invoice", "Invoice date", "Period from", "Period to",
            "Category", "Charge date", "Vendor", "Description", "Building total",
            "Share", "Your cost", "Service from", "Service to", "Flags"]
    ws.append(cols)
    _style_header(ws, 1, len(cols))
    line_flag_map = _line_flag_map(findings, items)
    for it in items:
        ty = tax_year(it["charge_date"], cfg.tax_start_day, cfg.tax_start_month)
        flags = line_flag_map.get(id(it), "")
        ws.append([ty, it["invoice_no"], fmt_date(it["invoice_date"]),
                   fmt_date(it["period_from"]), fmt_date(it["period_to"]),
                   it["category"], fmt_date(it["charge_date"]), it["vendor"],
                   it["description"], it["total_amount"], it["share"],
                   it["your_cost"], fmt_date(it["service_from"]),
                   fmt_date(it["service_to"]), flags])
        row = ws.max_row
        ws.cell(row=row, column=10).number_format = MONEY_FMT
        ws.cell(row=row, column=12).number_format = MONEY_FMT
        if it["your_cost"] < 0:
            for c in range(1, len(cols) + 1):
                ws.cell(row=row, column=c).fill = CREDIT_FILL
        elif flags:
            for c in range(1, len(cols) + 1):
                ws.cell(row=row, column=c).fill = ANOMALY_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    _autosize(ws, [9, 10, 12, 12, 12, 16, 12, 26, 50, 13, 7, 11, 12, 12, 40])

    # ---------- By Category ----------
    _matrix_sheet(wb, "By Category", items, order, invoices, key="category")
    # ---------- By Vendor ----------
    _matrix_sheet(wb, "By Vendor", items, order, invoices, key="vendor")

    # ---------- By Tax Year ----------
    ws = wb.create_sheet("By Tax Year")
    tys = sorted({tax_year(it["charge_date"], cfg.tax_start_day,
                           cfg.tax_start_month) for it in items})
    cats = sorted({it["category"] for it in items})
    ws.append(["Category"] + tys + ["Total"])
    _style_header(ws, 1, len(tys) + 2)
    grid = defaultdict(lambda: defaultdict(float))
    for it in items:
        grid[it["category"]][tax_year(it["charge_date"], cfg.tax_start_day,
                                     cfg.tax_start_month)] += it["your_cost"]
    for cat in cats:
        rowvals = [round(grid[cat].get(ty, 0.0), 2) for ty in tys]
        ws.append([cat] + rowvals + [round(sum(rowvals), 2)])
        for c in range(2, len(tys) + 3):
            ws.cell(row=ws.max_row, column=c).number_format = MONEY_FMT
    trow = ws.max_row + 1
    ws.cell(row=trow, column=1, value="TOTAL").font = Font(bold=True)
    for j, ty in enumerate(tys, start=2):
        v = round(sum(grid[cat].get(ty, 0.0) for cat in cats), 2)
        cell = ws.cell(row=trow, column=j, value=v)
        cell.number_format = MONEY_FMT
        cell.font = Font(bold=True)
    gt = ws.cell(row=trow, column=len(tys) + 2,
                 value=round(sum(it["your_cost"] for it in items), 2))
    gt.number_format = MONEY_FMT
    gt.font = Font(bold=True)
    for c in range(1, len(tys) + 3):
        ws.cell(row=trow, column=c).fill = TOTAL_FILL
    ws.freeze_panes = "B2"
    _autosize(ws, [22] + [12] * (len(tys) + 1))

    # ---------- Anomalies ----------
    ws = wb.create_sheet("Anomalies")
    cols = ["Severity", "Type", "Invoice", "Category", "Vendor", "Amount", "Detail"]
    ws.append(cols)
    _style_header(ws, 1, len(cols))
    sev_order = {"high": 0, "medium": 1, "low": 2}
    for f in sorted(findings, key=lambda x: (sev_order.get(x["severity"], 3), x["type"])):
        ws.append([f["severity"], f["type"], f["invoice_no"], f["category"],
                   f["vendor"], f["amount"], f["detail"]])
        row = ws.max_row
        if f["amount"] is not None:
            ws.cell(row=row, column=6).number_format = MONEY_FMT
        fill = SEVERITY_FILL.get(f["severity"])
        if fill:
            ws.cell(row=row, column=1).fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    _autosize(ws, [9, 30, 10, 16, 26, 11, 80])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _line_flag_map(findings, items):
    """Best-effort: tag line items whose description/invoice is named in a
    line-level finding so the Line Items sheet can highlight them."""
    out = {}
    snippet_types = {
        "Charge date outside invoice period",
        "Stale service period",
        "Possible duplicate line (same invoice)",
    }
    for it in items:
        tags = []
        for f in findings:
            if f["invoice_no"] != it["invoice_no"]:
                continue
            ftype = f["type"]
            if ftype in snippet_types:
                snippet = it["description"][:50]
                if snippet and snippet[:30] in f["detail"]:
                    tags.append(ftype)
            elif ftype == "Cross-invoice duplicate charge":
                sf, st = it.get("service_from"), it.get("service_to")
                if (f.get("vendor") == it["vendor"]
                        and sf and fmt_date(sf) in f["detail"]):
                    tags.append(ftype)
        if tags:
            out[id(it)] = "; ".join(sorted(set(tags)))
    return out


def _matrix_sheet(wb, title, items, order, invoices, key):
    ws = wb.create_sheet(title)
    labels = [f"{no}\n{fmt_date(invoices.get(no, {}).get('period_to'))}" for no in order]
    ws.append([key.capitalize()] + labels + ["Total"])
    _style_header(ws, 1, len(order) + 2)
    grid = defaultdict(lambda: defaultdict(float))
    for it in items:
        grid[it[key]][it["invoice_no"]] += it["your_cost"]
    rows = sorted(grid.keys())
    for name in rows:
        vals = [round(grid[name].get(no, 0.0), 2) for no in order]
        ws.append([name] + vals + [round(sum(vals), 2)])
        for c in range(2, len(order) + 3):
            ws.cell(row=ws.max_row, column=c).number_format = MONEY_FMT
    trow = ws.max_row + 1
    ws.cell(row=trow, column=1, value="TOTAL").font = Font(bold=True)
    for j, no in enumerate(order, start=2):
        v = round(sum(grid[name].get(no, 0.0) for name in rows), 2)
        cell = ws.cell(row=trow, column=j, value=v)
        cell.number_format = MONEY_FMT
        cell.font = Font(bold=True)
    gt = ws.cell(row=trow, column=len(order) + 2,
                 value=round(sum(it["your_cost"] for it in items), 2))
    gt.number_format = MONEY_FMT
    gt.font = Font(bold=True)
    for c in range(1, len(order) + 3):
        ws.cell(row=trow, column=c).fill = TOTAL_FILL
    ws.freeze_panes = "B2"
    _autosize(ws, [30] + [14] * (len(order) + 1))


# --------------------------------------------------------------------------
# Markdown report
# --------------------------------------------------------------------------
def write_invoice_report(target_no, items, invoices, order, findings, cfg, out_path: Path):
    """Per-invoice, point-in-time anomaly report for `target_no`.

    `order` ends with `target_no`; `items` covers only invoices up to it;
    `findings` are those attributable to `target_no`."""
    m = invoices.get(target_no, {})
    cur_items = [it for it in items if it["invoice_no"] == target_no]
    priors = order[:-1]

    lines = []
    lines.append(f"# Invoice {target_no} — anomaly report\n")
    lines.append(f"- Invoice date: {fmt_date(m.get('invoice_date'))}")
    lines.append(
        f"- Charge period: {fmt_date(m.get('period_from'))} to "
        f"{fmt_date(m.get('period_to'))}"
    )
    lines.append(
        f"- Tax year (by period end): "
        f"{tax_year(m.get('period_to'), cfg.tax_start_day, cfg.tax_start_month)}\n"
    )

    s = round(sum(i["your_cost"] for i in cur_items), 2)
    stated = m.get("invoice_total")
    if stated is not None:
        ok = "reconciles" if abs(s - stated) <= 0.01 else "**DOES NOT RECONCILE**"
        lines.append(
            f"**Reconciliation:** line items sum to {s:.2f}; stated invoice total "
            f"{stated:.2f} — {ok}.\n"
        )
    else:
        lines.append(f"**Reconciliation:** line items sum to {s:.2f}.\n")

    lines.append("## Cost by category (this invoice)\n")
    lines.append("| Category | Your cost |")
    lines.append("|---|---:|")
    cat_t = defaultdict(float)
    for it in cur_items:
        cat_t[it["category"]] += it["your_cost"]
    for cat in sorted(cat_t, key=lambda c: -cat_t[c]):
        lines.append(f"| {cat} | {round(cat_t[cat], 2):.2f} |")
    lines.append(f"| **Total** | **{s:.2f}** |")
    lines.append("")

    lines.append("## Cost by source / vendor (this invoice)\n")
    lines.append("| Vendor | Your cost |")
    lines.append("|---|---:|")
    ven_t = defaultdict(float)
    for it in cur_items:
        ven_t[it["vendor"]] += it["your_cost"]
    for v in sorted(ven_t, key=lambda x: -ven_t[x]):
        lines.append(f"| {v} | {round(ven_t[v], 2):.2f} |")
    lines.append("")

    sev_order = {"high": 0, "medium": 1, "low": 2}
    n_high = sum(1 for f in findings if f["severity"] == "high")
    n_med = sum(1 for f in findings if f["severity"] == "medium")
    n_low = sum(1 for f in findings if f["severity"] == "low")
    lines.append(
        f"## Anomalies: {len(findings)} ({n_high} high, {n_med} medium, {n_low} low)\n"
    )
    if priors:
        lines.append(f"*Compared against {len(priors)} earlier invoice(s): "
                     f"{', '.join(priors)}.*\n")
    else:
        lines.append("*Earliest invoice on record — no prior period to compare "
                     "against, so only within-invoice checks ran.*\n")

    if not findings:
        lines.append("No anomalies detected.\n")
    else:
        by_type = defaultdict(list)
        for f in findings:
            by_type[f["type"]].append(f)
        for atype in sorted(by_type, key=lambda t: min(sev_order.get(x["severity"], 3)
                                                        for x in by_type[t])):
            group = sorted(by_type[atype], key=lambda x: sev_order.get(x["severity"], 3))
            lines.append(f"### {atype} ({len(group)})\n")
            for f in group:
                lines.append(f"- **{f['severity'].upper()}** {f['detail']}")
            lines.append("")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input-dir", default=".", help="folder with the two CSVs")
    ap.add_argument("--output-dir", default=None, help="defaults to input-dir")
    ap.add_argument("--items", default="line_items.csv")
    ap.add_argument("--invoices", default="invoices.csv")
    ap.add_argument("--tax-start-day", type=int, default=6)
    ap.add_argument("--tax-start-month", type=int, default=4)
    ap.add_argument("--high-ratio", type=float, default=1.5)
    ap.add_argument("--low-ratio", type=float, default=0.5)
    ap.add_argument("--material", type=float, default=15.0,
                    help="min absolute £ difference for a value anomaly")
    cfg = ap.parse_args()

    in_dir = Path(cfg.input_dir)
    out_dir = Path(cfg.output_dir) if cfg.output_dir else in_dir

    invoices = load_invoices(in_dir / cfg.invoices)
    items = load_items(in_dir / cfg.items)
    order = sorted(invoices, key=lambda n: (invoices[n]["period_from"] or date.min,
                                            invoices[n]["invoice_date"] or date.min))

    # One point-in-time report per invoice: each invoice is analysed as the
    # "current" period, seeing only the invoices chronologically before it.
    all_findings = []
    reports = []
    for idx, no in enumerate(order):
        sub_order = order[: idx + 1]
        sub_set = set(sub_order)
        sub_items = [it for it in items if it["invoice_no"] in sub_set]
        f = detect_anomalies(sub_items, invoices, sub_order, cfg)
        rep = out_dir / f"anomaly_report_{no}.md"
        write_invoice_report(no, sub_items, invoices, sub_order, f, cfg, rep)
        all_findings.extend(f)
        reports.append((no, rep, f))

    xlsx = out_dir / "cost_tracker.xlsx"
    build_workbook(items, invoices, order, all_findings, cfg, xlsx)

    grand = round(sum(it["your_cost"] for it in items), 2)
    print(f"Invoices: {len(order)}  Line items: {len(items)}  "
          f"Total your-cost: {grand:.2f}")
    print(f"Wrote: {xlsx}")
    print("Per-invoice reports (newest first):")
    for no, rep, f in reversed(reports):
        nh = sum(1 for x in f if x["severity"] == "high")
        print(f"  {rep.name}: {len(f)} anomalies ({nh} high)")


if __name__ == "__main__":
    main()
